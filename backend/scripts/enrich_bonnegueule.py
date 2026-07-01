"""Enrichit les échelles de Vetements.xlsx avec des marques FR intermédiaires.

Curation informée BonneGueule : les listes actuelles sont déjà riches (japonais/
anglo + luxe) mais sans marque FR mid-tier. On insère ces marques juste après la
marque d'entrée. Master gitignoré : la modif est locale (backup + re-sync).

Usage (depuis backend/) :
    uv run python -m scripts.enrich_bonnegueule
"""
from __future__ import annotations


def _norm(s: str | None) -> str:
    return " ".join(str(s or "").strip().casefold().split())


def insert_after_entry(echelle: list[str], brands: list[str]) -> list[str]:
    """Insère `brands` juste après `echelle[0]`, dédupliquées (casse/espaces).

    N'ajoute que des marques absentes de `echelle` et non dupliquées entre elles ;
    ordre préservé. Ne retire jamais → len(résultat) >= len(echelle).
    """
    def _dedup(items: list[str], already: set[str]) -> list[str]:
        out: list[str] = []
        seen = set(already)
        for b in items:
            k = _norm(b)
            if not k or k in seen:
                continue
            seen.add(k)
            out.append(b)
        return out

    if not echelle:
        return _dedup(brands, set())
    new = _dedup(brands, {_norm(x) for x in echelle})
    return [echelle[0]] + new + echelle[1:]


# Type objectif -> marques FR intermédiaires à insérer (cf. spec).
ENRICHMENT: dict[str, list[str]] = {
    # Hauts
    "T-shirts": ["Loom", "Asphalte", "Le Minor"],
    "Polos": ["Asphalte", "Le Minor"],
    "Chemises": ["Asphalte", "Officine Générale", "De Bonne Facture"],
    "Débardeurs": ["Le Minor", "Saint James"],
    "Pulls": ["Officine Générale", "De Bonne Facture", "Maison Montagut"],
    "Sweats": ["Asphalte", "Maison Labiche"],
    "Gilets": ["Officine Générale", "De Bonne Facture"],
    "Overshirts": ["Asphalte", "De Bonne Facture"],
    # Bas
    "Jeans": ["Asphalte", "1083", "Ateliers de Nîmes"],
    "Pantalons chino": ["Asphalte", "Officine Générale", "De Bonne Facture"],
    "Pantalons habillés": ["Officine Générale", "De Fursac", "Husbands"],
    "Shorts": ["Asphalte"],
    "Pantalons en velours": ["Officine Générale", "De Bonne Facture"],
    # Vestes / manteaux
    "Vestes légères": ["Officine Générale", "De Bonne Facture", "Harmony"],
    "Blazers": ["Officine Générale", "De Fursac", "Husbands"],
    "Manteaux": ["Officine Générale", "Harmony", "Éditions M.R"],
    # Costume
    "Vestes de costume": ["De Fursac", "Husbands", "Samson"],
    "Pantalons de costume": ["De Fursac", "Husbands"],
    "Gilets de costume": ["De Fursac", "Husbands"],
    "Smokings": ["De Fursac", "Husbands"],
    # Sous-vêtements
    "Boxers": ["Le Slip Français"],
    "Slips": ["Le Slip Français"],
    "Maillots de corps": ["Le Slip Français"],
    "Chaussettes": ["Bleuforêt", "Labonal", "Royalties"],
    # Chaussures
    "Bottines": ["Jules & Jenn", "Anthology Paris"],
    "Chaussures de ville": ["Bexley", "Jacques & Déclercq", "Markowski"],
    # Accessoires
    "Ceintures": ["JOSEPH BONNIE", "Bleu de Chauffe"],
    "Cravates": ["Le Colonel Moutarde", "Cinabre"],
    "Nœuds papillon": ["Le Colonel Moutarde", "Cinabre"],
    "Foulards": ["Cinabre"],
    "Lunettes de soleil": ["Jimmy Fairly", "Ateliers Loden"],
    "Sacs à dos": ["Bleu de Chauffe", "Bonastre", "Côme"],
    "Sacs de voyage": ["Bleu de Chauffe", "Bonastre"],
    "Portefeuilles": ["Bleu de Chauffe", "JOSEPH BONNIE"],
    # Technique
    "Maillots techniques": ["Circle Sportswear"],
    "Shorts techniques": ["Circle Sportswear"],
}


def _echelle_from_row(ws, row: int) -> list[str]:
    """Marques d'une ligne (colonnes C.. = 3..), non vides."""
    out: list[str] = []
    for col in range(3, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None and str(v).strip():
            out.append(str(v).strip())
    return out


def main() -> dict:
    import datetime as dt
    import shutil

    import openpyxl

    from app.core.config import settings

    path = settings.imports_dir / "Vetements.xlsx"
    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = path.with_name(f"Vetements.backup-{ts}.xlsx")
    shutil.copy2(path, backup)

    wb = openpyxl.load_workbook(path)  # writable (pas read_only)
    ws = wb.worksheets[0]

    found_types: set[str] = set()
    enriched = 0
    for row in range(2, ws.max_row + 1):
        nom_cell = ws.cell(row=row, column=1).value
        if not nom_cell or not str(nom_cell).strip():
            continue
        nom = str(nom_cell).strip()
        found_types.add(nom)
        if nom not in ENRICHMENT:
            continue
        old = _echelle_from_row(ws, row)
        new = insert_after_entry(old, ENRICHMENT[nom])
        if new == old:
            continue
        # insert_after_entry n'ajoute que -> len(new) >= len(old), pas de nettoyage
        for i, brand in enumerate(new):
            ws.cell(row=row, column=3 + i, value=brand)
        enriched += 1
        print(f"{nom}: +{len(new) - len(old)}  {old} -> {new}")

    not_found = sorted(t for t in ENRICHMENT if t not in found_types)
    wb.save(path)
    print(f"types enrichis: {enriched} / {len(ENRICHMENT)} ; backup: {backup.name} ; non trouvés: {not_found}")
    return {"enriched": enriched, "backup": str(backup), "not_found": not_found}


if __name__ == "__main__":
    main()

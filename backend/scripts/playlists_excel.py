"""Outil autonome : Excel <-> playlists musicales (DeepSeek) -> .m3u8.

Flux : l'Excel est la source de vérité ÉDITABLE.
  1. export   : DB MusicTrack -> playlists.xlsx (métadonnées + 1 colonne par
                playlist, vides) + colonnes techniques (chemin, duree_sec).
  2. classify : remplit les colonnes True/False via DeepSeek (par lots).
  3. (tu corriges à la main dans Excel.)
  4. build    : génère un .m3u8 par playlist (UTF-8 BOM) depuis l'Excel -> zip.

Usage (depuis backend/) :
  uv run python -m scripts.playlists_excel export   [out.xlsx]
  uv run python -m scripts.playlists_excel classify [out.xlsx]
  uv run python -m scripts.playlists_excel build    [out.xlsx] [dossier_sortie]
"""
from __future__ import annotations

import json
import re
import sys
import time
import zipfile
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from app.core.config import settings
from app.core.db import engine
from app.models.musique import MusicTrack
from app.services.musique.playlists import safe_filename, to_m3u8

# Catalogue des 14 playlists (label affiché -> description pour guider DeepSeek).
PLAYLISTS: list[tuple[str, str]] = [
    ("café pour le petit dep", "léger, doux, agréable au réveil / petit-déjeuner"),
    ("coworking/travail/detente", "rythmé mais non distrayant, fond de travail/concentration"),
    ("soirée ( francophone )", "festif, dansant, chansons francophones"),
    ("soirée ( internationale )", "festif, dansant, hits internationaux"),
    ("amour/love/sex", "romantique, sensuel, intime"),
    ("chanson francaise", "chanson française : variété et auteurs-compositeurs francophones"),
    ("tristesse", "triste, mélancolique, doux-amer, introspectif"),
    ("sport/gym", "entraînant, tempo élevé, motivation sportive"),
    ("good vibes", "positif, ensoleillé, joyeux, feel-good"),
    ("chill", "détendu, posé, douce ambiance (lo-fi, downtempo)"),
    ("rock", "rock au sens large : classic rock, alternatif, indie, rock FR et international"),
    ("pop", "pop grand public, mélodies accrocheuses"),
    ("dance/EDM", "électronique dansante : house, EDM, club, dance"),
    ("rap", "rap / hip-hop, français et international"),
]
PLAYLIST_LABELS = [label for label, _ in PLAYLISTS]
META_COLS = ["titre", "album", "auteur", "co-auteur"]
TECH_COLS = ["chemin", "duree_sec"]
HEADERS = META_COLS + PLAYLIST_LABELS + TECH_COLS
BATCH_SIZE = 20

_DEFAULT_XLSX = "playlists.xlsx"

# Séparateurs de featuring courants (on évite ',' et 'x' : trop de faux positifs).
# '/' et '&' : avec espaces optionnels ; feat/ft/featuring : entourés d'espaces,
# point optionnel consommé.
_SEP = re.compile(r"\s*/\s*|\s*&\s*|\s+(?:feat|ft|featuring)\.?\s+", re.IGNORECASE)


def split_artists(artist: str, title: str) -> tuple[str, str]:
    """(auteur principal, co-auteurs) au mieux depuis le champ artiste + le titre."""
    feats: list[str] = []
    m = re.search(r"\bfeat\.?\s+([^)\]]+)\)?", title or "", re.IGNORECASE)
    if m:
        feats.append(m.group(1).strip())
    parts = [p.strip() for p in _SEP.split(artist or "") if p.strip()]
    primary = parts[0] if parts else (artist or "").strip()
    rest = parts[1:] + feats
    seen: set[str] = set()
    co = [c for c in rest if not (c.lower() in seen or seen.add(c.lower()))]
    return primary, ", ".join(co)


def _build_prompt(rows: list[dict]) -> str:
    lignes = "\n".join(f"- {label} : {desc}" for label, desc in PLAYLISTS)
    morceaux = "\n".join(
        f"{i}. titre={r['titre']}, album={r['album']}, artiste={r['auteur']}"
        + (f" feat {r['co-auteur']}" if r["co-auteur"] else "")
        for i, r in enumerate(rows, start=1)
    )
    labels = ", ".join(f'"{label}"' for label in PLAYLIST_LABELS)
    return (
        "Tu classes des morceaux de musique par playlist. Playlists possibles :\n"
        f"{lignes}\n\n"
        "Pour chaque morceau, indique TOUTES les playlists réellement adaptées "
        "(0 à 5). Inclus les genres ET les ambiances qui collent.\n\n"
        f"Morceaux :\n{morceaux}\n\n"
        "Réponds UNIQUEMENT par un objet JSON de la forme :\n"
        '{"resultats": [{"index": <numéro>, "playlists": [<playlists>]}]}\n'
        f"Chaque playlist doit être exactement l'une de : {labels}."
    )


def _parse_resultats(payload: dict, n: int) -> list[list[str]]:
    """payload DeepSeek -> pour chaque morceau (index 1..n), la liste de labels valides."""
    valid = set(PLAYLIST_LABELS)
    out: list[list[str]] = [[] for _ in range(n)]
    for item in payload.get("resultats", []):
        i = item.get("index", 0) - 1
        if not 0 <= i < n:
            continue
        vues: list[str] = []
        for label in item.get("playlists", []):
            if label in valid and label not in vues:
                vues.append(label)
        out[i] = vues
    return out


def classify_batch(rows: list[dict], *, _post=None) -> list[list[str]]:
    post = _post or httpx.post
    resp = post(
        f"{settings.deepseek_base_url.rstrip('/')}/chat/completions",
        headers={
            "Authorization": f"Bearer {settings.deepseek_api_key}",
            "Content-Type": "application/json",
        },
        json={
            "model": settings.musique_deepseek_model,
            "messages": [
                {"role": "system", "content": "Tu réponds uniquement par du JSON valide."},
                {"role": "user", "content": _build_prompt(rows)},
            ],
            "response_format": {"type": "json_object"},
            "temperature": 0,
            "stream": False,
        },
        timeout=120.0,
    )
    resp.raise_for_status()
    text = resp.json()["choices"][0]["message"]["content"]
    return _parse_resultats(json.loads(text), len(rows))


# --------------------------------------------------------------------------- #
# Commandes
# --------------------------------------------------------------------------- #
def cmd_export(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "playlists"
    ws.append(HEADERS)
    n = 0
    with Session(engine) as s:
        for t in s.exec(select(MusicTrack).order_by(MusicTrack.path)).all():
            auteur, co = split_artists(t.artist, t.title)
            ws.append([t.title, t.album, auteur, co]
                      + [None] * len(PLAYLIST_LABELS)
                      + [t.path, t.duree_sec])
            n += 1
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"export OK : {n} morceaux -> {path}")


def _read_rows(ws):
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2):
        rows.append({h: r[idx[h]].value for h in headers} | {"_cells": r, "_idx": idx})
    return rows


def cmd_classify(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    rows = _read_rows(ws)
    pl_col = {label: rows[0]["_idx"][label] for label in PLAYLIST_LABELS} if rows else {}
    done = 0
    for i in range(0, len(rows), BATCH_SIZE):
        lot = rows[i:i + BATCH_SIZE]
        payload = [{"titre": r["titre"] or "", "album": r["album"] or "",
                    "auteur": r["auteur"] or "", "co-auteur": r["co-auteur"] or ""} for r in lot]
        res = None
        for attempt in range(2):
            try:
                res = classify_batch(payload)
            except Exception as e:  # noqa: BLE001
                print(f"  lot {i // BATCH_SIZE} essai {attempt} échec: {type(e).__name__}: {e}")
                res = None
            if res and any(res):
                break
            time.sleep(0.5)
        if not res:
            res = [[] for _ in lot]
        for r, labels in zip(lot, res):
            labset = set(labels)
            for label in PLAYLIST_LABELS:
                r["_cells"][pl_col[label]].value = label in labset
        wb.save(path)  # progrès persistant à chaque lot
        done += len(lot)
        if (i // BATCH_SIZE) % 5 == 0:
            ne = sum(1 for x in res if x)
            print(f"  ... lot {i // BATCH_SIZE}: {ne}/{len(lot)} non vides (cumul {done}/{len(rows)})")
        time.sleep(0.2)
    print(f"classify OK : {done} morceaux écrits -> {path}")


def cmd_build(path: Path, out_dir: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    rows = _read_rows(ws)
    out_dir.mkdir(parents=True, exist_ok=True)
    zip_path = out_dir / "playlists-musique.zip"
    counts = {}
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for label in PLAYLIST_LABELS:
            tracks = []
            for r in rows:
                val = r[label]
                if val is True or (isinstance(val, str) and val.strip().lower() in ("true", "vrai", "1", "x")):
                    tracks.append({"path": r["chemin"], "artist": r["auteur"],
                                   "title": r["titre"], "duree_sec": r["duree_sec"]})
            fname = f"{safe_filename(label)}.m3u8"
            data = to_m3u8(tracks, titre=label)
            (out_dir / fname).write_bytes(data)
            zf.writestr(fname, data)
            counts[label] = len(tracks)
    print(f"build OK -> {zip_path}")
    for label in PLAYLIST_LABELS:
        print(f"  {label}: {counts[label]}")


def main(argv: list[str]) -> int:
    if not argv or argv[0] not in {"export", "classify", "build"}:
        print(__doc__)
        return 2
    cmd = argv[0]
    path = Path(argv[1]) if len(argv) > 1 else Path(_DEFAULT_XLSX)
    if cmd == "export":
        cmd_export(path)
    elif cmd == "classify":
        cmd_classify(path)
    else:
        out_dir = Path(argv[2]) if len(argv) > 2 else Path("playlists_m3u8")
        cmd_build(path, out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

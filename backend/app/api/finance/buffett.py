"""Sous-routeur Finance : analyses Buffett (runs, progression, export, optimisation)."""
from __future__ import annotations

import io
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from app.core.db import get_session
from app.api.schemas_finance import (
    BuffettRunOut, BuffettRunDetailOut, BuffettProgressOut,
)
from app.models.finance import BuffettRun, BuffettRunResult

from app.core.rate_limit import rate_limit

logger = logging.getLogger(__name__)
router = APIRouter()

# Rate limit partagé des analyses Buffett coûteuses (#193) : 5 lancements/min/IP.
_analysis_rl = rate_limit(max_calls=5, window_s=60, name="finance_analysis")


@router.get("/buffett/runs", response_model=list[BuffettRunOut])
def buffett_runs(limit: int = 12, session: Session = Depends(get_session)):
    stmt = (select(BuffettRun)
            .order_by(BuffettRun.run_date.desc())
            .limit(limit))
    return list(session.exec(stmt).all())


@router.get("/buffett/latest", response_model=Optional[BuffettRunOut])
def buffett_latest(session: Session = Depends(get_session)):
    stmt = (select(BuffettRun)
            .where(BuffettRun.statut == "termine")
            .order_by(BuffettRun.run_date.desc()))
    return session.exec(stmt).first()


@router.get("/buffett/progress", response_model=BuffettProgressOut)
def buffett_progress(session: Session = Depends(get_session)):
    from app.services.finance.scheduler_stub import is_analysis_running
    active = is_analysis_running()

    # Si aucune analyse ne tourne reellement dans ce process, un run encore
    # "en_cours" en base est en fait interrompu (programme ferme) -> on le marque
    # immediatement resumable pour debloquer le bouton "Reprendre".
    if not active:
        stuck = session.exec(
            select(BuffettRun).where(
                BuffettRun.statut.in_(["en_cours", "interrompu"])  # type: ignore[attr-defined]
            )
        ).all()
        changed = False
        for sr in stuck:
            # Analyse complète (100 %) mais process fermé avant le marquage final :
            # on la termine au lieu de boucler indéfiniment sur "Reprendre".
            if (sr.n_tickers_total or 0) > 0 and (sr.n_tickers_analyzed or 0) >= sr.n_tickers_total:
                sr.statut = "termine"
                sr.erreur = None
                session.add(sr)
                changed = True
            elif sr.statut == "en_cours":
                sr.statut = "interrompu"
                sr.erreur = "Process interrompu (relancez pour reprendre)"
                session.add(sr)
                changed = True
        if changed:
            session.commit()

    run = session.exec(
        select(BuffettRun)
        .where(BuffettRun.statut.in_(["en_cours", "interrompu"]))  # type: ignore[attr-defined]
        .order_by(BuffettRun.created_at.desc())
    ).first()
    if run:
        from app.services.finance.buffett.rate_limiter import active_paused_until
        return BuffettProgressOut(
            run_id=run.id, statut=run.statut, active=active,
            progress_pct=run.progress_pct or 0.0,
            n_done=run.n_tickers_analyzed, n_total=run.n_tickers_total,
            paused_until=active_paused_until() if active else None,
        )
    return BuffettProgressOut(run_id=None, statut="idle", active=False, progress_pct=0.0)


@router.get("/buffett/runs/{run_id}", response_model=BuffettRunDetailOut)
def buffett_run_detail(run_id: int, session: Session = Depends(get_session)):
    run = session.get(BuffettRun, run_id)
    if not run:
        raise HTTPException(404, f"Run {run_id} introuvable")
    top = list(session.exec(
        select(BuffettRunResult)
        .where(BuffettRunResult.run_id == run_id)
        .order_by(BuffettRunResult.chance_moat.desc())
        .limit(50)
    ).all())
    alloc = list(session.exec(
        select(BuffettRunResult)
        .where(BuffettRunResult.run_id == run_id)
        .where(BuffettRunResult.allocation_pct.isnot(None))
        .order_by(BuffettRunResult.allocation_pct.desc())
    ).all())
    return BuffettRunDetailOut(run=run, top_results=top, allocation_cible=alloc)


@router.delete("/buffett/runs/{run_id}", status_code=204)
def buffett_run_delete(run_id: int, session: Session = Depends(get_session)):
    """Supprime un run Buffett (et ses résultats) — ex. analyse bloquée."""
    from app.services.finance.buffett.reporting import delete_run
    if not delete_run(session, run_id):
        raise HTTPException(404, f"Run {run_id} introuvable")


@router.get("/buffett/runs/{run_id}/export")
def buffett_run_export(run_id: int, session: Session = Depends(get_session)):
    """Exporte les resultats d'un run Buffett en Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    run = session.get(BuffettRun, run_id)
    if not run:
        raise HTTPException(404, f"Run {run_id} introuvable")

    results = list(session.exec(
        select(BuffettRunResult)
        .where(BuffettRunResult.run_id == run_id)
        .order_by(BuffettRunResult.chance_moat.desc())
    ).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultats MOAT"

    ws.append([f"Analyse Buffett - Run #{run_id} - {run.run_date}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([
        f"Statut : {run.statut}",
        f"Tickers analyses : {run.n_tickers_analyzed or 0}/{run.n_tickers_total or 0}",
        f"Duree : {run.duree_sec or 0:.0f}s",
    ])
    ws.append([])

    headers = ["Ticker", "Nom", "Secteur", "Pays", "Score MOAT", "Alloc. cible (%)", "Broker cible", "Achat"]
    header_row = 4
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    green_fill = PatternFill("solid", fgColor="D6F5D6")
    blue_fill = PatternFill("solid", fgColor="D6E8F5")
    for r in results:
        score_val = r.chance_moat
        row = [
            r.ticker,
            r.nom or "",
            r.secteur or "",
            r.pays or "",
            round(score_val, 2) if score_val is not None else None,
            round(r.allocation_pct, 2) if r.allocation_pct is not None else None,
            r.broker_cible or "",
            "oui" if r.achat else "",
        ]
        ws.append(row)
        if score_val and score_val >= 200:
            fill = blue_fill  # ETF
        elif score_val and score_val >= 80:
            fill = green_fill  # Eligible
        else:
            fill = None
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = fill

    col_widths = [12, 32, 20, 18, 12, 16, 16, 7]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    alloc_list = [r for r in results if r.allocation_pct]
    if alloc_list:
        ws2 = wb.create_sheet("Allocation cible")
        ws2.append(["Ticker", "Nom", "Broker cible", "Alloc. cible (%)"])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        for r in sorted(alloc_list, key=lambda x: -(x.allocation_pct or 0)):
            ws2.append([r.ticker, r.nom or "", r.broker_cible or "",
                        round(r.allocation_pct, 2) if r.allocation_pct else None])
        for i, w in enumerate([12, 32, 16, 16], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"buffett_run_{run_id}_{run.run_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/buffett/runs/{run_id}/export.csv")
def buffett_run_export_csv(run_id: int, session: Session = Depends(get_session)):
    """Exporte les résultats d'un run Buffett en CSV (sans dépendance)."""
    import csv
    import io as _io

    run = session.get(BuffettRun, run_id)
    if not run:
        raise HTTPException(404, f"Run {run_id} introuvable")

    results = list(session.exec(
        select(BuffettRunResult)
        .where(BuffettRunResult.run_id == run_id)
        .order_by(BuffettRunResult.chance_moat.desc())
    ).all())

    buf = _io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    writer.writerow([
        "Ticker", "Nom", "Secteur", "Pays", "Score MOAT", "Achat",
        "Allocation cible (%)", "Broker cible", "PER", "Prix",
    ])
    for r in results:
        writer.writerow([
            r.ticker, r.nom or "", r.secteur or "", r.pays or "",
            round(r.chance_moat, 2) if r.chance_moat is not None else "",
            "oui" if r.achat else "",
            round(r.allocation_pct, 2) if r.allocation_pct is not None else "",
            r.broker_cible or "",
            round(r.per, 2) if r.per else "",
            round(r.prix, 2) if r.prix else "",
        ])
    buf.seek(0)
    filename = f"buffett_run_{run_id}_{run.run_date}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/buffett/breakdown/{ticker}")
def buffett_breakdown(ticker: str):
    """Détail du score Buffett d'un titre par critère (marge, ROE, dette…)."""
    from app.services.finance.buffett.runner import analyze_single_ticker
    from app.services.finance.buffett.scoring_pure import score_breakdown

    result = analyze_single_ticker(ticker.upper().strip())
    if result is None:
        raise HTTPException(404, f"Analyse impossible pour {ticker} (données absentes/invalides).")
    score, metrics = result
    ratios = metrics.get("ratios_recents") or {}
    return {
        "ticker": ticker.upper(),
        "score": score,
        "secteur": metrics.get("Secteur"),
        "criteres": score_breakdown(ratios),
    }


@router.get("/buffett/taux-obligataires")
def buffett_bond_yields():
    """Taux obligataires utilisés par le critère d'achat (rafraîchis en direct).

    Confirme que les taux ne sont plus statiques : les pays dans `live` sont
    récupérés à la journée (cache), les autres utilisent le repli statique.
    """
    from app.services.finance.buffett.bond_yields import (
        SERIES_BY_COUNTRY,
        STATIC_BOND_YIELDS,
        get_bond_yields,
    )

    taux = get_bond_yields(defaults=dict(STATIC_BOND_YIELDS))
    live = {
        pays: taux[pays]
        for pays in SERIES_BY_COUNTRY
        if pays in taux and taux[pays] != STATIC_BOND_YIELDS.get(pays)
    }
    return {
        "taux": taux,
        "live": live,                       # pays (G7) effectivement rafraîchis aujourd'hui
        "sources": SERIES_BY_COUNTRY,       # pays -> série FRED
        "repli": STATIC_BOND_YIELDS,
    }


@router.get("/backtest")
def backtest_allocation(periode: str = "2y", session: Session = Depends(get_session)):
    """Backtest buy-and-hold de l'allocation cible du dernier run Buffett terminé.

    Renvoie {dates, equity (base 100), rendement_pct, n_points, tickers}.
    """
    run = session.exec(
        select(BuffettRun).where(BuffettRun.statut == "termine").order_by(BuffettRun.run_date.desc())
    ).first()
    if not run:
        raise HTTPException(404, "Aucun run Buffett terminé")

    rows = list(session.exec(
        select(BuffettRunResult)
        .where(BuffettRunResult.run_id == run.id)
        .where(BuffettRunResult.allocation_pct.isnot(None))  # type: ignore[attr-defined]
    ).all())
    weights = {r.ticker: float(r.allocation_pct or 0) for r in rows if (r.allocation_pct or 0) > 0}
    if not weights:
        return {"dates": [], "equity": [], "rendement_pct": 0.0, "n_points": 0, "tickers": []}

    from app.services.finance.backtest import simulate_allocation
    dates: list[str] = []
    prices: dict[str, list[float]] = {}
    try:
        import yfinance as yf
        import pandas as pd
        from app.services.finance.yf_session import yf_session
        t_list = list(weights.keys())
        raw = yf.download(t_list, period=periode, interval="1d", progress=False, group_by="ticker", session=yf_session())
        if not raw.empty:
            if len(t_list) == 1:
                cd = raw["Close"].to_frame(); cd.columns = t_list
            else:
                cd = pd.DataFrame({
                    t: raw[t]["Close"] for t in t_list
                    if t in raw.columns.get_level_values(0)
                })
            cd = cd.dropna(how="all").ffill().dropna()
            dates = [d.strftime("%Y-%m-%d") for d in cd.index]
            prices = {t: [float(x) for x in cd[t].tolist()] for t in cd.columns}
    except Exception as exc:
        logger.warning("Backtest download: %s", exc)

    sim = simulate_allocation(prices, weights)
    dates = dates[: sim["n_points"]]
    return {
        "dates": dates,
        "equity": sim["equity"],
        "rendement_pct": sim["rendement_pct"],
        "n_points": sim["n_points"],
        "tickers": list(prices.keys()),
    }


# --- Bouton 1 : Analyser tous les tickers ---

@router.post("/buffett/run", status_code=202, dependencies=[Depends(_analysis_rl)])
def buffett_run_start(
    background_tasks: BackgroundTasks,
    csv_path: Optional[str] = None,
    session: Session = Depends(get_session),
):
    """Lance ou REPREND l'analyse de tous les tickers de tickers.csv.

    Si une analyse tourne deja dans ce process, l'appel est ignore. Sinon le job
    reprend automatiquement le dernier run interrompu (sans refaire les tickers
    deja analyses, sauvegardes un a un) ou en cree un nouveau.
    """
    from app.services.finance.scheduler_stub import job_monthly_buffett, is_analysis_running
    if is_analysis_running():
        return {"message": "Analyse deja en cours", "status": "running"}
    background_tasks.add_task(job_monthly_buffett, csv_path)
    return {"message": "Analyse demarree (reprise si un run etait interrompu)", "status": "accepted"}


# --- Bouton 2 : Analyser un ticker precis ---

@router.post("/buffett/analyze-ticker", status_code=200, dependencies=[Depends(_analysis_rl)])
async def buffett_analyze_ticker(ticker: str):
    """Analyse un seul ticker et retourne score + metrics immediatement."""
    from app.services.finance.buffett.runner import analyze_single_ticker
    result = analyze_single_ticker(ticker.upper().strip())
    if result is None:
        raise HTTPException(
            404,
            f"Impossible d'analyser {ticker} : donnees absentes, trop fraiches, ou ticker invalide.",
        )
    score, metrics = result
    return {"ticker": ticker.upper(), "score": score, "metrics": metrics}


# --- Bouton 3 : Creer le portefeuille optimal (Differential Evolution) ---

@router.post("/portfolio/create", status_code=202, dependencies=[Depends(_analysis_rl)])
def portfolio_create(
    background_tasks: BackgroundTasks,
    min_score: float = 80.0,
    session: Session = Depends(get_session),
):
    """Filtre les eligibles, re-verifie les scores, optimise avec DE."""
    latest_run = session.exec(
        select(BuffettRun)
        .where(BuffettRun.statut == "termine")
        .order_by(BuffettRun.run_date.desc())
    ).first()
    if latest_run is None:
        raise HTTPException(404, "Aucun run Buffett termine. Lancer d'abord l'analyse complete.")

    def _run_portfolio_creation(run_id: int, min_score_val: float) -> None:
        from app.services.finance.buffett.config import Config
        from app.services.finance.buffett.optimizer import optimize_portfolio_de, prepare_optimization
        from app.services.finance.buffett.allocation import discretize_allocation, latest_prices
        from app.services.finance.buffett.broker_availability import merge_broker_columns
        from app.services.finance.buffett.reporting import update_allocations
        from app.services.finance.buffett.dedup import deduplicate_tickers
        from app.core.db import engine
        from sqlmodel import Session as S, select as sel
        import pandas as pd
        import numpy as np
        import yfinance as yf

        Config.load_params()

        with S(engine) as sess:
            # Tous les candidats: score >= seuil OU ETF (score=200) OU Achat=True
            all_rows = list(sess.exec(
                sel(BuffettRunResult).where(BuffettRunResult.run_id == run_id)
            ).all())

        candidates = {
            r.ticker: r for r in all_rows
            if (r.chance_moat or 0) >= min_score_val
            or (r.chance_moat or 0) >= 200  # ETF
            or r.achat
        }
        if not candidates:
            logger.warning("[portfolio_create] Aucun candidat eligible.")
            return

        logger.info(f"[portfolio_create] {len(candidates)} candidats (depuis le run en DB)...")

        # On NE re-télécharge PAS chaque candidat en live : sous rate-limit
        # yfinance, ces appels échouaient tous -> 0 ticker valide -> DE sur liste
        # vide -> Sharpe -inf. On réutilise les scores/indicateurs déjà calculés
        # par le run (en DB). Le seul appel réseau restant est le download groupé
        # des cours pour la matrice de covariance (impersoné via yf_session).
        forced = [t.upper() for t in Config.FORCED_BUY_TICKERS]
        verified: dict[str, tuple[float, dict]] = {}
        for ticker, r in candidates.items():
            score = float(r.chance_moat or 0)
            is_etf = (r.secteur == "ETF") or score >= 200
            is_forced = ticker.upper() in forced
            eligible = score >= min_score_val or is_etf or is_forced
            if eligible and r.achat:
                verified[ticker] = (score, {
                    "Nom": r.nom or "", "Secteur": r.secteur or "",
                    "Volume": r.volume or 0, "Achat": True,
                })

        logger.info(f"[portfolio_create] {len(verified)} tickers valides -> optimisation DE...")
        if not verified:
            logger.warning("[portfolio_create] Aucun ticker valide -> optimisation annulee.")
            return

        t_list = list(verified.keys())
        ticker_col = "Ticker Yahoo Finance"
        try:
            from app.services.finance.yf_session import yf_session
            raw = yf.download(t_list, period="5y", interval="1d", progress=False, group_by="ticker", session=yf_session())
            if raw.empty:
                return
            if len(t_list) == 1:
                cd = raw["Close"].to_frame()
                cd.columns = t_list
            else:
                cd = pd.DataFrame({
                    t: raw[t]["Close"] for t in t_list
                    if t in raw.columns.get_level_values(0)
                })
            cd = cd.dropna(axis=1, thresh=len(cd) * 0.01).ffill()
            rets = cd.pct_change().dropna().clip(-0.5, 0.5)

            df_m = pd.DataFrame([{
                ticker_col: t,
                "Nom": verified[t][1].get("Nom", ""),
                "Secteur": verified[t][1].get("Secteur", ""),
                "Volume": verified[t][1].get("Volume", 0),
                "Chance MOAT": verified[t][0],
                "Achat": True,
            } for t in t_list if t in rets.columns])

            # Disponibilite par broker depuis ToutBroker.xlsx (sinon tout dispo)
            df_m = merge_broker_columns(df_m, ticker_col)
            rets = deduplicate_tickers(rets, df_m, ticker_col)
            t_opt = list(rets.columns)
            if len(t_opt) < 2:
                logger.warning(
                    f"[portfolio_create] {len(t_opt)} ticker(s) avec historique de cours "
                    "exploitable -> optimisation impossible (besoin d'au moins 2)."
                )
                return
            mat_access, active_b = prepare_optimization(t_opt, df_m)
            weights, sharpe = optimize_portfolio_de(t_opt, rets, mat_access, active_b)

            total_cap = sum(Config.BUDGET_BROKERS.values())
            # Actions entieres (hors Trading212) / pies (Trading212), a partir des prix
            prices = latest_prices(cd, t_opt)
            alloc = discretize_allocation(t_opt, weights, active_b, prices, total_cap)
            with S(engine) as sess:
                update_allocations(sess, run_id, alloc, reset=True)
            # Écrire le poids (%) de chaque action dans ToutBroker.xlsx (#1)
            try:
                from app.services.finance.buffett.broker_availability import update_broker_file_weights
                n_w = update_broker_file_weights(alloc)
                logger.info(f"[portfolio_create] {n_w} poids ecrits dans ToutBroker.xlsx")
            except Exception as e:
                logger.error(f"[portfolio_create] Ecriture Poids: {e}")
            n_alloc = len({a["Ticker"] for a in alloc})
            logger.info(f"[portfolio_create] Sharpe={sharpe:.3f}, {n_alloc} tickers alloues.")
        except Exception as e:
            logger.error(f"[portfolio_create] Erreur optimisation: {e}")

    background_tasks.add_task(_run_portfolio_creation, latest_run.id, min_score)
    return {
        "message": f"Creation portefeuille lancee (run #{latest_run.id}, seuil={min_score})",
        "status": "accepted",
        "run_id": latest_run.id,
    }

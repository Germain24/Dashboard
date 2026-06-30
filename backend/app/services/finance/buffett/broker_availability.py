"""Disponibilité des actions par broker, lue depuis ToutBroker.xlsx.

Le fichier (par défaut ``data/imports/ToutBroker.xlsx``) contient une ligne par
action avec des colonnes de disponibilité par broker — ``Tradding 212``,
``Bourse Direct``, ``Bourse Direct 2``, ``IBKR`` — valant 1 (disponible) / 0
(indisponible). On fusionne ces colonnes dans le DataFrame passé à l'optimiseur
afin que celui-ci n'alloue une action que sur les brokers où elle est achetable.

Ticker inconnu dans le fichier → considéré disponible partout (comportement legacy).
"""

from __future__ import annotations

import os
import re

from .config import Config


def _clean(v) -> str:
    return "".join(filter(str.isalnum, str(v).upper()))


def _alpha(v: str) -> str:
    return "".join(c for c in v if c.isalpha())


def _trailing_num(v: str) -> str:
    m = re.search(r"(\d+)$", v)
    return m.group(1) if m else ""


def find_broker_file() -> str | None:
    """Cherche ToutBroker.xlsx dans plusieurs emplacements plausibles."""
    rel = os.path.join("data", "imports", "Finances", "tableur", "ToutBroker.xlsx")
    candidates = [
        Config.BROKER_FILE,                 # data/imports/Finances/tableur/ToutBroker.xlsx (absolu)
        rel,
        os.path.join("..", rel),
    ]
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return None


def load_broker_table():
    """Charge ToutBroker.xlsx en DataFrame, ou None s'il est introuvable."""
    path = find_broker_file()
    if not path:
        return None
    try:
        import pandas as pd
        return pd.read_excel(path)
    except Exception as e:
        print(f"[broker_availability] Lecture {path}: {e}")
        return None


def _secteur1_col(columns) -> str | None:
    """Trouve la colonne 'Secteur 1' (tolère espaces/casse)."""
    for c in columns:
        if str(c).strip().lower() == "secteur 1":
            return c
    return None


def _norm_ticker(v) -> str:
    s = str(v).strip().upper()
    return "" if s in ("", "NAN", "NONE") else s


def _compute_etf_tickers(df, ticker_col: str) -> set[str]:
    if df is None or getattr(df, "empty", True):
        return set()
    tcol = _find_ticker_col(df.columns, ticker_col)
    scol = _secteur1_col(df.columns)
    if tcol is None or scol is None:
        return set()
    out: set[str] = set()
    for t, s in zip(df[tcol], df[scol]):
        if str(s).strip().upper() == "ETF":
            tt = _norm_ticker(t)
            if tt:
                out.add(tt)
    return out


_ETF_CACHE: set[str] | None = None
_UNIVERSE_CACHE: set[str] | None = None


def reset_etf_cache() -> None:
    """Invalide les caches ETF/univers (à appeler après modif de ToutBroker.xlsx)."""
    global _ETF_CACHE, _UNIVERSE_CACHE
    _ETF_CACHE = None
    _UNIVERSE_CACHE = None


def load_etf_tickers(df=None, ticker_col: str = "Ticker Yahoo Finance") -> set[str]:
    """Tickers (MAJ) dont 'Secteur 1' == 'ETF' dans ToutBroker.xlsx.

    Source AUTORITAIRE de la classification ETF (score=200). ``df`` explicite =
    pas de cache (tests) ; sinon résultat mémoïsé (relire l'Excel à chaque ticker
    serait prohibitif). Utiliser ``reset_etf_cache()`` après une écriture du fichier.
    """
    global _ETF_CACHE
    if df is not None:
        return _compute_etf_tickers(df, ticker_col)
    if _ETF_CACHE is None:
        _ETF_CACHE = _compute_etf_tickers(load_broker_table(), ticker_col)
    return _ETF_CACHE


def _compute_universe(df, ticker_col: str) -> set[str]:
    if df is None or getattr(df, "empty", True):
        return set()
    tcol = _find_ticker_col(df.columns, ticker_col)
    if tcol is None:
        return set()
    return {tt for t in df[tcol] if (tt := _norm_ticker(t))}


def load_broker_universe(df=None, ticker_col: str = "Ticker Yahoo Finance") -> set[str]:
    """Tous les tickers (MAJ) présents dans ToutBroker.xlsx (univers curé)."""
    global _UNIVERSE_CACHE
    if df is not None:
        return _compute_universe(df, ticker_col)
    if _UNIVERSE_CACHE is None:
        _UNIVERSE_CACHE = _compute_universe(load_broker_table(), ticker_col)
    return _UNIVERSE_CACHE


def _cell_state(v) -> bool | None:
    """Interprète une cellule de disponibilité broker : True / False / None (vide).

    None = cellule vide / valeur inattendue → « inconnu » (ne compte pas comme Faux).
    """
    try:
        import pandas as pd
        if v is None or pd.isna(v):
            return None
    except (TypeError, ValueError):
        if v is None:
            return None
    s = str(v).strip().lower()
    if s == "" or s == "nan":
        return None
    if s in ("1", "1.0", "true", "vrai", "oui", "yes", "x", "v"):
        return True
    if s in ("0", "0.0", "false", "faux", "non", "no"):
        return False
    return None  # valeur non reconnue → prudence : analyser


def broker_excluded_tickers(ticker_col: str = "Ticker Yahoo Finance") -> set[str]:
    """Tickers (MAJ) à IGNORER : présents dans ToutBroker.xlsx ET dont TOUTES les
    colonnes broker sont explicitement Faux. Une cellule vide ne compte pas comme
    Faux (la ligne reste analysée). Tout le reste (absent du fichier, au moins un
    broker Vrai, ou au moins une cellule vide) est analysé.

    Vide si le fichier est introuvable / sans colonne ticker / sans colonne broker
    reconnue → l'appelant n'exclut alors rien (analyse tout, garde-fou).
    """
    df = load_broker_table()
    if df is None:
        return set()
    tcol = _find_ticker_col(df.columns, ticker_col)
    if tcol is None:
        return set()
    broker_cols: list = []
    for broker in Config.BUDGET_BROKERS:
        c = _match_broker_column(broker, df.columns)
        if c is not None and c not in broker_cols:
            broker_cols.append(c)
    if not broker_cols:
        return set()

    excluded: set[str] = set()
    for _, row in df.iterrows():
        ticker = str(row[tcol]).strip()
        if not ticker or ticker.lower() == "nan":
            continue
        states = [_cell_state(row[c]) for c in broker_cols]
        # Exclure uniquement si TOUTES les colonnes broker sont explicitement Faux
        # (aucune Vrai, aucune vide).
        if states and all(s is False for s in states):
            excluded.add(ticker.upper())
    return excluded


def _find_ticker_col(columns, ticker_col: str) -> str | None:
    if ticker_col in columns:
        return ticker_col
    for c in columns:
        if "ticker" in str(c).lower():
            return c
    return None


def _match_broker_column(broker_name: str, columns) -> str | None:
    """Associe un broker de Config (ex: 'Trading212') à sa colonne dans ToutBroker
    (ex: 'Tradding 212'), tolérant l'orthographe et les espaces.
    """
    cb, cb_num = _clean(broker_name), _trailing_num(broker_name)
    cb_alpha = _alpha(cb)
    best = None
    for c in columns:
        cc = _clean(c)
        if cb == cc:
            return c  # correspondance exacte
        cc_num, cc_alpha = _trailing_num(cc), _alpha(cc)
        if cc_num != cb_num:
            continue
        # même numéro de fin + préfixes alpha qui se ressemblent (4 premiers car.)
        if cb_alpha[:4] and cb_alpha[:4] == cc_alpha[:4]:
            best = c
    return best


# Attribut BuffettRunResult -> colonne ToutBroker.xlsx. Seules les colonnes
# déjà présentes dans le fichier sont écrites (on n'invente pas de colonnes).
_RESULT_TO_BROKER_COL = {
    "chance_moat": "Chance MOAT",
    "achat": "Achat",
    "nom": "Nom",
    "pays": "Pays",
    "secteur": "Secteur",
    "prix": "Prix",
    "eps": "EPS",
    "per": "PER",
    "croissance": "Croissance",
    "peg": "PEG",
    "volume": "Volume",
}


def _save_main_sheet(df, path: str) -> None:
    """Écrit ``df`` dans la 1re feuille de ToutBroker en PRÉSERVANT les autres
    feuilles (``ETF_Defensif``, ``ETF_Pays``…). Sans autres feuilles -> écriture simple.
    """
    import pandas as pd
    main, has_others = None, False
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        main = wb.sheetnames[0]
        has_others = len(wb.sheetnames) > 1
        wb.close()
    except Exception:
        pass
    if has_others and main:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            df.to_excel(w, sheet_name=main, index=False)
    else:
        df.to_excel(path, index=False)


def update_broker_file_scores(rows, path: str | None = None,
                              ticker_col: str = "Ticker Yahoo Finance") -> int:
    """Écrit les scores/indicateurs de l'analyse dans ToutBroker.xlsx (upsert par ticker).

    ``rows`` : itérable d'objets type ``BuffettRunResult`` (attributs ``ticker``,
    ``chance_moat``, ``achat``, ``nom``, ``pays``, ``secteur``, ``prix``, ``eps``,
    ``per``, ``croissance``, ``peg``, ``volume``).

    - Ticker déjà présent -> met à jour ses cellules (Chance MOAT, Achat, indicateurs).
    - Ticker absent -> ajoute une nouvelle ligne.
    - Les autres colonnes (disponibilité broker incluse) sont **préservées**.

    Retourne le nombre de tickers traités. Sans fichier -> 0 (aucune écriture).
    """
    path = path or find_broker_file()
    if not path:
        return 0
    try:
        import pandas as pd

        df = pd.read_excel(path)
        tcol = _find_ticker_col(df.columns, ticker_col) or ticker_col
        if tcol not in df.columns:
            return 0

        # Colonnes du fichier qu'on sait remplir.
        writable = {attr: col for attr, col in _RESULT_TO_BROKER_COL.items() if col in df.columns}

        # Excel relit souvent les colonnes "entières" (8.0, 12.0…) en int64 ;
        # passer en object évite un rejet de dtype lors de l'écriture d'un float.
        for col in set(writable.values()):
            df[col] = df[col].astype(object)

        # Index ticker (nettoyé) -> position de ligne (première occurrence).
        index: dict[str, int] = {}
        for i, k in df[tcol].astype(str).str.strip().items():
            index.setdefault(k, i)

        new_rows: list[dict] = []
        n = 0
        for r in rows:
            tk = str(getattr(r, "ticker", "") or "").strip()
            if not tk:
                continue
            payload = {}
            for attr, col in writable.items():
                val = getattr(r, attr, None)
                if attr == "achat":
                    val = bool(val)
                payload[col] = val
            if tk in index:
                for col, val in payload.items():
                    df.at[index[tk], col] = val
            else:
                row = {tcol: tk}
                row.update(payload)
                new_rows.append(row)
            n += 1

        if new_rows:
            df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

        _save_main_sheet(df, path)
        return n
    except Exception as e:
        print(f"[broker_availability] Écriture scores ToutBroker: {e}")
        return 0


def aggregate_weights(alloc: list[dict]) -> dict[str, float]:
    """Pur : somme du « Poids total (%) » par ticker depuis une allocation.

    ``alloc`` est la sortie de ``discretize_allocation`` (une ligne par couple
    ticker/broker). Le poids d'une action = somme de ses lignes.
    """
    out: dict[str, float] = {}
    for row in alloc or []:
        tk = str(row.get("Ticker", "") or "").strip()
        if not tk:
            continue
        out[tk] = out.get(tk, 0.0) + float(row.get("Poids total (%)", 0) or 0)
    return {t: round(p, 4) for t, p in out.items()}


def update_broker_file_weights(alloc: list[dict], path: str | None = None,
                               ticker_col: str = "Ticker Yahoo Finance",
                               weight_col: str = "Poids") -> int:
    """Écrit le pourcentage d'investissement par action dans la colonne ``Poids``.

    Agrège l'allocation par ticker (somme des poids par broker), remet toute la
    colonne ``Poids`` à 0, puis inscrit le poids de chaque action allouée.
    Crée la colonne si elle n'existe pas. Retourne le nombre de tickers écrits.
    """
    weights = aggregate_weights(alloc)
    path = path or find_broker_file()
    if not path:
        return 0
    try:
        import pandas as pd

        df = pd.read_excel(path)
        tcol = _find_ticker_col(df.columns, ticker_col) or ticker_col
        if tcol not in df.columns:
            return 0

        if weight_col not in df.columns:
            df[weight_col] = 0.0
        df[weight_col] = 0.0  # reset (les non-alloués passent à 0, pas de valeur obsolète)

        index: dict[str, int] = {}
        for i, k in df[tcol].astype(str).str.strip().items():
            index.setdefault(k, i)

        n = 0
        for tk, pct in weights.items():
            if tk in index:
                df.at[index[tk], weight_col] = pct
                n += 1
        _save_main_sheet(df, path)
        return n
    except Exception as e:
        print(f"[broker_availability] Écriture Poids ToutBroker: {e}")
        return 0


def merge_broker_columns(df_m, ticker_col: str = "Ticker Yahoo Finance"):
    """Ajoute à ``df_m`` une colonne de disponibilité par broker de Config.BUDGET_BROKERS,
    renommée exactement comme la clé Config pour que l'optimiseur la retrouve.

    Sans fichier ToutBroker.xlsx, ``df_m`` est renvoyé inchangé (tout disponible).
    """
    tbl = load_broker_table()
    if tbl is None or getattr(tbl, "empty", True):
        return df_m
    try:
        tcol = _find_ticker_col(tbl.columns, ticker_col)
        if tcol is None:
            return df_m
        sub = tbl.copy()
        sub[tcol] = sub[tcol].astype(str).str.strip()
        sub = sub.drop_duplicates(subset=[tcol], keep="last")
        lookup = sub.set_index(tcol)

        out = df_m.copy()
        keys = list(df_m[ticker_col].astype(str).str.strip())
        for broker in Config.BUDGET_BROKERS:
            col = _match_broker_column(broker, tbl.columns)
            if not col:
                continue
            series = lookup[col]
            out[broker] = [series.get(k, None) for k in keys]
        # Colonne ISIN (dédup ETF par identité exacte), si présente dans ToutBroker.
        isin_col = next((c for c in tbl.columns if str(c).strip().upper() == "ISIN"), None)
        if isin_col is not None:
            series = lookup[isin_col]
            out["ISIN"] = [series.get(k, None) for k in keys]
        return out
    except Exception as e:
        print(f"[broker_availability] Fusion colonnes broker: {e}")
        return df_m
